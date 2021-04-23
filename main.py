from bs4 import BeautifulSoup
from nltk import probability
from scipy.sparse.construct import random
import seaborn as sns
from openpyxl import Workbook, load_workbook
import pandas as pd
import requests
import re
import os
import json
import numpy as np
from collections import Counter
import argparse
import math

import warnings
warnings.filterwarnings(action = 'ignore')

from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.manifold import TSNE, MDS
import nltk
nltk.download(['punkt', 'wordnet'])
from nltk.tokenize import word_tokenize, MWETokenizer
from nltk.stem import WordNetLemmatizer
from gensim.corpora import Dictionary
from gensim.similarities import MatrixSimilarity
from gensim.models import TfidfModel, LdaMulticore
import scipy

## Utility function(s) and global variables:
def keywordSubstitute(keyword):
	if keyword == 'DoS attack':
		keyword = 'denial of service'
	elif keyword == 'malicious bot':
		keyword = 'bot$'
	elif keyword == 'encryption':
		keyword = '^encrypt'
	elif keyword == 'targeted threat':
		keyword = 'target.. threat'

	return keyword

## Problem 1
def loadKeywordsIntoDict():
	"""Return the workbook, active sheet in the workbook, keyword dictionary 

	@description - Extracts keywords from the excel sheet and transforms them into a nested dictionary
	@param - None
	
	"""
	keywords_workbook = load_workbook('./keywords.xlsx')
	active_sheet = keywords_workbook.active
	keywords = {}

	for x in range(2, active_sheet.max_row + 1):
		keywords[active_sheet.cell(row=x, column=1).value] = []

	return keywords_workbook, active_sheet, keywords

def extractNewsInfo(article, sourceName): 
	"""Return boolean value if article is a news article, URL of the article

	@description - Gets type and url link of news article from HTML
	@param - HTML code of one article, name of article source
	
	"""
	if sourceName == 'BBC':
		soup2 = BeautifulSoup(str(article), 'lxml')
		article_type = (soup2.find_all('span', 'ecn1o5v0'))[1].text
		article_link = (soup2.find('a', 'e1f5wbog5', href=True))['href']
		
		if article_type == 'News':
			return True, article_link
		
		return False, None

	else:
		soup2 = BeautifulSoup(str(article), 'lxml')
		article_link = (soup2.find('a', href=True))['href']
		
		return article_link

def isRelevant(article_link, current_keyword, keyword_list):
	resp = requests.get(article_link)
	soup = BeautifulSoup(resp.text, 'html.parser')
	try:
		article_content = soup.find('article', 'e1nh2i2l0').find_all('p') # gets list of sentences in article
	except:
		try:
			article_content = soup.find('body').find_all('table')[7].find_all('p') # same logic as line 106 but used on older BBC news articles
		except:
			return False

	for k in keyword_list:
		k = keywordSubstitute(k)
		if (k != current_keyword) and (re.search(k, str(article_content))):
			return True

def getArticlesBBC(keywords):
	"""Return null

	@description - Gets all articles from keyword search via BBC News
	@param - Empty keyword dictionary
	
	"""
	print('Begin Data Collection ----------------------------------')
	for keyword in keywords:
		print(f'Looking for articles related to {keyword} from BBC')
		page_count = 1

		if keyword == 'DoS attack':
			query_keyword = 'denial of service attack'
		else:
			query_keyword = keyword

		while len(keywords[keyword]) <= 100:
			params = {'q' : query_keyword, 'page' : page_count} # BBC News URL Format: https://www.bbc.co.uk/search?q={keyword}&page={number}
			url = 'https://www.bbc.co.uk/search'
			try:
				resp = requests.get(url, params=params)
				if resp.status_code != 200:
					break
			except:
				break
			page_count += 1 # assists with traversing through pages
			soup = BeautifulSoup(resp.text, 'html.parser')
			articles = soup.find('ul', class_="e1y4nx260").find_all('li')

			if len(articles) == 0:
				break
			
			for article in articles:
				isNews, article_link = extractNewsInfo(article, 'BBC')
				# print('Checking relevancy of {}'.format(article_link))
				# and isRelevant(article_link, keyword, keywords.keys())
				if isNews:
					keywords[keyword].append(article_link)

		if len(keywords[keyword]) > 100:
			keywords[keyword] = keywords[keyword][:100]

def getArticlesCyberpost(keywords):
	"""Return a pointer to indicate the start of the links for articles from this source

	@description - Gets all articles from keyword search via Cyberpost to supplement the lack of articles for some keywords in BBC News
	@param - Empty keyword dictionary
	
	"""
	start_ptr = []

	for keyword in keywords:
		start_ptr.append(len(keywords[keyword]))
		print(f'Looking for articles related to {keyword} from Cyberpost')
		page_count = 1
		
		while len(keywords[keyword]) <= 100:
			url = 'https://thecyberpost.com/page/'+str(page_count)+'/'

			try:
				resp = requests.get(url, params={'s' : keyword})
				if resp.status_code != 200:
					break
			except:
				break

			page_count += 1
			soup = BeautifulSoup(resp.text, 'html.parser')
			articles = soup.find('div', class_="td-ss-main-content").find_all('div', class_='td_module_16')
			if len(articles) == 0:
				break
			
			for article in articles:
				article_link = extractNewsInfo(article, 'Cyberpost')
				keywords[keyword].append(article_link)

		if len(keywords[keyword]) > 100:
			keywords[keyword] = keywords[keyword][:100]

	print('Data Collection Completed ----------------------------------')

	return start_ptr

## Problem 2
class TextPreprocessing:
	punctuations = [',', '.', ';', ':', '"', "'", '`', '|', '/', '``', ' ', '']
	sw = [' ','','\n','%', '``', 'page', 'break', 'ourselves', 'hers', 'between', 'yourself', 'but', 'again', 
			'there', 'about', 'once', 'during', 'out', 'very', 'having', 'with', 'they', 'own', 'an', 'be', 'some', 
			'for', 'do', 'its', 'yours', 'such', 'into', 'of', 'most', 'itself', 'other', 'off', 
			'is', 's', 'am', 'or', 'who', 'as', 'from', 'him', 'each', 'the', 'themselves', 
			'until', 'below', 'are', 'we', 'these', 'your', 'his', 'through', 'don', 'nor', 
			'me', 'were', 'her', 'more', 'himself', 'this', 'down', 'should', 'our', 'their', 
			'while', 'above', 'both', 'up', 'to', 'ours', 'had', 'she', 'all', 'no', 'when', 
			'at', 'any', 'before', 'them', 'same', 'and', 'been', 'have', 'in', 'will', 'on', 
			'does', 'yourselves', 'then', 'that', 'because', 'what', 'over', 'why', 'so', 'can', 
			'did', 'not', 'now', 'under', 'he', 'you', 'herself', 'has', 'just', 'where', 'too', 
			'only', 'myself', 'which', 'those', 'i', 'after', 'few', 'whom', 't', 'being', 'if', 
			'theirs', 'my', 'against', 'a', 'by', 'doing', 'it', 'how', 'further', 'was', 'here', 'than'
		]
	def __init__(self, special_words):
		self.lemT = WordNetLemmatizer()
		self.mweT = MWETokenizer([tuple(w.lower().split(' ')) for w in special_words if len(w.split(' ')) > 1])
		self.mweT.add_mwe(('denial', 'service')) 
		self.mweT.add_mwe(('denial', 'of', 'service')) 
		self.mweT.add_mwe(('(DDOS)', 'attack'))
	def __call__(self, text):
		lemmatization = [self.lemT.lemmatize(token.lower()) for token in word_tokenize(text) if token not in self.punctuations]
		# processed_lemmatization = self.mweT.tokenize(initial_lemmatization)
		return [i for i in lemmatization if i not in self.sw]

def saveArticle(contentList, keyword, special_words):
	"""Return null

	@description - Saves all text from articles into named files
	@param - List containing text, keyword which the program is currently looking at, index counter
	
	"""
	preprocessor = TextPreprocessing(special_words)
	filename = f'{keyword}_articles.txt'
	with open(f'./Article_Contents/{filename}', 'w', encoding='utf-8') as f:
		for article in contentList:
			if str(article) != '--page-break--':
				for line in article:
					sentence = line.getText(strip=True)
					tokenized_sentence = preprocessor(str(sentence))
					sentence = ' '.join(tokenized_sentence)
					f.write(' \n '+sentence+' \n ')

			else:
				f.write(' \n '+str(article)+' \n ')

	f.close()

def processArticle(keywords, start_ptr):
	"""Return null
	
	@description - Gets all articles from keyword search via BBC News
	@param - Empty keyword dictionary, list of pointers to indicate change in source
	
	"""
	print('Begin Data Preprocessing --------------------------------------')
	index = 0

	if os.path.exists('Article_Contents') == False:
		os.mkdir('Article_Contents')

	for keyword in keywords:
		article_content = []
		print(f'Extracting and processing content from articles of {keyword}')

		for counter, link in enumerate(keywords[keyword]): # visiting each article link
			
			resp = requests.get(link)
			soup3 = BeautifulSoup(resp.text, 'html.parser')

			if counter < start_ptr[index]: # checks if we are looking at BBC or Cyberpost articles

				try:
					article_content.append(soup3.find('article', 'e1nh2i2l0').find_all('p')) # gets list of sentences in article
				except:
					try:
						article_content.append(soup3.find('body').find_all('table')[7].find_all('p')) # same logic as line 106 but used on older BBC news articles
					except:
						continue

				try:
					subheadings = soup3.find('article', 'e1nh2i2l0').find_all('h2')
					article_content.append(subheadings)
				except:
					pass

				article_content.append('--page-break--')

			else:

				try:
					article_content.append(soup3.find('div', class_='bialty-container').find_all('p'))
				except:
					continue

				article_content.append('--page-break--')

		saveArticle(article_content, keyword, list(keywords.keys()))

		index += 1

	print('Data Preprocessing Completed ----------------------------------')

## Problem 3

# Algorithm 1 : Cosine Similarity using TF-IDF--------------------------
def create_dataframe(matrix, tokens):

    doc_names = [f'doc_{i+1}' for i, _ in enumerate(matrix)]
    df = pd.DataFrame(data=matrix, index=doc_names, columns=tokens)
    return df

def vectorizeDocuments(word1, word2):
	sw = [' ','','\n', ' --page-break--']

	with open(f'./Article_Contents/{word1}_articles.txt', 'r', encoding='utf-8') as f1:
		article1 = f1.read()
	f1.close()

	with open(f'./Article_Contents/{word2}_articles.txt', 'r', encoding='utf-8') as f2:
		article2 = f2.read()
	f2.close()

	dataToVectorize = [article1, article2]
	tfidf = TfidfVectorizer(stop_words=sw)
	vector = tfidf.fit_transform(dataToVectorize)
	tokenized_words = tfidf.get_feature_names()
	vectorToArray = vector.toarray()

	df = create_dataframe(vectorToArray, tokenized_words)
	return vectorToArray
	
def calculateCosineSimilarity(word1, word2):
	combinedVector = vectorizeDocuments(word1, word2)
	vectorA, vectorB = np.array(combinedVector[0]), np.array(combinedVector[1])
	dotproduct = np.dot(vectorA, vectorB)
	cosSim = round(dotproduct / (np.linalg.norm(vectorA)*np.linalg.norm(vectorB)), 3)

	# print(f"similarity between {word1} and {word2}: {cosSim}")
	return cosSim, combinedVector

# ----------------------------------------------------------------------

# Algorithm 2: Jenson Shannon Divergence -------------------------------

class JensonShannonDist:
	def probability2(self, words, wordsToCheck, n=20):
		words = {key:value for (key,value) in words.items() if key in wordsToCheck}
		data = np.array(list(words.values()))
		h, bin_edges = np.histogram(data, n)
		prob = h / data.shape[0]
		return bin_edges, prob
	
	def probability(self, words, wordsToCheck):
		words = {key:value for (key,value) in words.items() if key in wordsToCheck}
		wordsToCheck = {key:value for (key,value) in wordsToCheck.items() if key in words}
		total_count = sum(words.values())
		sorted_dict = sorted(words.items())
		print(sorted_dict)
		keys = []
		probabilities = []
		
		for key, value in sorted_dict:
			prob = value / total_count
			probabilities.append(prob)
			keys.append(key)

		return keys, probabilities

	def compute_jensen_shannon_dist(self, query, matrix):
		"""
		method to compute the Jenson-Shannon Distance 
		between two probability distributions
		"""
		# p = query[None, :].T
		# q = matrix.T

		p = query
		q = matrix

		m = 0.5*(p + q)

		# Jensen Shannon Divergence
		divergence = 0.5*(scipy.stats.entropy(p, m) + scipy.stats.entropy(q, m))

		# Jensen Shannon Distance
		dist = np.sqrt(divergence)
		return dist

	def compare(self, train_words, test_words):
		p_keys, p_prob = self.probability(train_words, test_words)
		_, q_prob = self.probability(test_words, train_words)

		# be1, p_prob = self.probability2(train_words, test_words, 10)
		# be2, q_prob = self.probability2(test_words, train_words, 10)

		p = np.array(p_prob)
		q = np.array(q_prob)

		dist = self.compute_jensen_shannon_dist(p, q)
		# print(dist)
		return dist

# Algorithm 3: Latent Dirichlet Allocation -----------------------------

class LDA:
	def __init__(self, keywords):
		self.df = []
		self.training_corpus = []
		self.keywords = [k for k in keywords]
		self.bowDict = {}
		self.model = None
		self.tfidf = None
		self.simMatrix = None

	def readFiles(self, folderName):
		processed_docs = []
		for k in self.keywords:
			with open(f'{folderName}/{k}_articles.txt', 'r', encoding='utf-8') as f:
				processed_docs.append(f.read().split(' '))
			f.close()

		self.df = pd.DataFrame(np.array(processed_docs), columns=['document'])
		self.df['title'] = self.keywords

	def createBOW(self):
		bowDict = Dictionary(self.df['document'].values)
		bowDict.filter_extremes(no_below=2, no_above=0.35, keep_n=100000)
		corpus = [bowDict.doc2bow(doc) for doc in self.df['document']]
		return corpus, bowDict

	def generateTFIDF(self):
		self.training_corpus, self.bowDict = self.createBOW()
		self.tfidf = TfidfModel(self.training_corpus, smartirs='npu')
		tfidf_corpus = self.tfidf[self.training_corpus]
		self.model = LdaMulticore(tfidf_corpus, num_topics=10, id2word=self.bowDict, passes=2, workers=2)
		self.simMatrix = MatrixSimilarity(self.model[tfidf_corpus])

	def calculateSim(self, dist_matrix, debug='false'):
		for index, doc in enumerate(self.df['document']):
			keyword = self.df['title'][index]
			test_vec = self.bowDict.doc2bow(doc)
			tfidf_bow = self.tfidf[test_vec]
			vec_lsi = self.model[tfidf_bow]
			similarity = self.simMatrix[vec_lsi]

			if debug == 'true':
				print(f'----{keyword}----')
				for s in sorted(enumerate(similarity), key=lambda item: -item[1])[:10]:
					print(f"{self.df['title'].iloc[s[0]]} : {str(s[1])}")
				print('------------------------')

			dist_matrix[index] = [round(1 - s, 5) for s in similarity]
		np.fill_diagonal(dist_matrix, 0.)

	def calculateJS(self, dist_matrix, debug='false'):
		jensenshannon = JensonShannonDist()
		vectors = []
		for index, doc in enumerate(self.df['document']):
			keyword = self.df['title'][index]
			test_vec = self.bowDict.doc2bow(doc)
			tfidf_bow = self.tfidf[test_vec]
			vectors.append([prob for (_, prob) in self.model[tfidf_bow]])

			if debug == 'true':
				print('---------------{}---------------'.format(keyword))
				for index, score in sorted(self.model[tfidf_bow], key=lambda tup: -1*tup[1]):
					print('\nScore: {}\t \nTopic: {}'.format(score, self.model.print_topic(index, 10)))

		for pos1, x in enumerate(vectors):
			for pos2, y in enumerate(vectors):
				dist = round(jensenshannon.compute_jensen_shannon_dist(np.array(x), np.array(y)), 5)
				if math.isnan(float(dist)):
					dist = 0.00000
				dist_matrix[pos1][pos2] = dist
				if args.debug == 'true':
					print(f"Distance between {x} and {y}: {dist}")

# -----------------------------------------------------------------------

def getWordFreq(keyword):
	sw = [' ','','\n', ' --page-break--']
	with open(f'Article_Contents/{keyword}_articles.txt') as f:
		data = f.read().split(' ')
		data = [i for i in data if i not in sw]
	
	return dict(Counter(data))

def createDistanceSpreadsheet(args, keywords_workbook, active_sheet, keywords):
	"""Return null
	
	@description - Creates an excel file named 'distance.xlsx' which contains the keyword distances values
	@param - Excel workbook, active worksheet in the workbook, keyword dictionary
	
	"""
	print('Begin Distance Calculation ----------------------------------')
	keyword_dists = np.full((len(keywords), len(keywords)), -1.)
	keys = list(keywords.keys())
	wordVectors = []
	mode = args.mode

	if mode != 'cosine_tfidf':

		lda = LDA(keys)
		lda.readFiles('./Article_Contents')
		if lda.model == None:
			lda.generateTFIDF()
		
		if mode == 'cosine_lda':
			print('Using Latent Dirichlet Allocation with Cosine Similarity')
			lda.calculateSim(keyword_dists, args.debug)
		else:
			print('Using Latent Dirichlet Allocation with Janson Shannon Distance')
			lda.calculateJS(keyword_dists, args.debug)

	else:
		print('Using TF-IDF with Cosine Similarity')

		for _, word1 in enumerate(keys):
			if args.debug == 'true':
				print('-----------------------------------------')
			for word2 in keys:

				if keyword_dists[keys.index(word1)][keys.index(word2)] != -1:
					continue

				if word1 == word2:
					keyword_dists[keys.index(word1)][keys.index(word2)] = 0
					continue

				else:
					dist, wordVector = calculateCosineSimilarity(word1, word2)
					dist = round(1 - dist, 5) # cosine distance is 1 - {cosine similarity}
					if list(wordVector[0]) in wordVector:
						wordVectors.append(list(wordVector[0]))

					keyword_dists[keys.index(word1)][keys.index(word2)] = dist
					keyword_dists[keys.index(word2)][keys.index(word1)] = dist

					if args.debug == 'true':
						print(f"Distance between {word1} and {word2}: {dist}")
				
	row_pos = 2

	for col_num in range(2, len(keywords)+2):
		active_sheet.cell(row=1, column=col_num).value = active_sheet.cell(row=col_num, column=1).value
	
	for i in range (len(keyword_dists)):
		col_pos = 2
		for j in range (len(keyword_dists[i])):
			active_sheet.cell(row=row_pos, column=col_pos).value = keyword_dists[i][j]
			col_pos += 1
		row_pos += 1

	keywords_workbook.save('distance.xlsx')
	print('Distance Calculation Completed and Saved in distance.xlsx ----------------------------------')
	return wordVectors

## Problem 4
def visusalizeDistance(vecMatrix=[], model_type='cosine_tfidf'):
	
	"""Return null
	
	@description - Creates barplots to visualize the distances between every keyword
	@param - null
	
	"""
	print('Generating graph(s)')
	df = pd.read_excel('distance.xlsx', 'Sheet1', index_col=None, na_values=['NA'])
	df.set_index('Keywords', inplace=True) # allows for easy selection of rows by row title

	## Heatmap ----------------------------------------------------------------------------------------------------------
	# sns.utils.plt.figure(figsize=(9, 6))
	# sns.utils.plt.subplots_adjust(left=0.27, right=0.99, bottom=0.33)
	# diagram = sns.heatmap(df, annot=True, annot_kws={'size':10}, fmt='1', cmap='BuPu', linecolor='white', linewidths=0.5)
	# diagram.set_xticklabels(diagram.get_xticklabels(), rotation=45, horizontalalignment='right')
	# diagram.set(xlabel='Keywords', ylabel='Keywords', title='Distance Heatmap')
	# diagram = diagram.get_figure()
	# sns.utils.plt.savefig(f'Distance_diagram_{model_type}.png', format='png', dpi=100)
	# sns.utils.plt.show()
	## ------------------------------------------------------------------------------------------------------------------

	## Clustermap ----------------------------------------------------------------------------------------------------------
	# diagram = sns.clustermap(df, figsize=(7,6), cbar_pos=(0.9, 0.5, 0.05, 0.3))
	# sns.utils.plt.savefig('Distance_diagram.png', format='png', dpi=100)
	# sns.utils.plt.show()

	## ------------------------------------------------------------------------------------------------------------------

	## Scatterplot --------------------------------------------------------------------------------------------------------
	
	# fig = sns.utils.plt.figure(figsize=(16,10))
	# fig.suptitle('Scatterplot for Keyword')
	# tsne = TSNE(n_components=2, verbose=1, perplexity=20, n_iter=300, metric='precomputed', learning_rate=200)
	# distMatrix = np.array(df.to_numpy())
	# result = tsne.fit_transform(distMatrix)
	# sns.scatterplot(result[:,0], result[:,1], hue=df.columns, legend='full', palette=sns.color_palette('bright', 10))
	# sns.utils.plt.savefig(f'Scatterplot_distance_tsne_{model_type}.png', format='png', dpi=100)
	# sns.utils.plt.show()

	fig = sns.utils.plt.figure(figsize=(16,10))
	fig.suptitle('Scatterplot for Keyword')
	mds = MDS(n_components=2, verbose=1, random_state=None, dissimilarity='precomputed', max_iter=500)
	distMatrix = np.array(df.to_numpy())
	result = mds.fit_transform(distMatrix)
	sns.scatterplot(result[:,0], result[:,1], hue=df.columns, legend='full', palette=sns.color_palette('bright', 10))
	# sns.utils.plt.savefig(f'Scatterplot_distance_mds_{model_type}.png', format='png', dpi=100)
	sns.utils.plt.show()

	# ---------------------------------------------------------------------------------------------------------------------

def argparser():
	parser = argparse.ArgumentParser()
	parser.add_argument('-m', '--mode', 
						help='Select method of calculating keyword distance',
						type=str,
						choices=['cosine_lda', 'cosine_tfidf', 'js_lda'],
						default='cosine_tfidf'
						)
	parser.add_argument('-d', '--debug',
						help='Enter "yes" for program to print results',
						type=str,
						choices=['true', 'false'],
						default='false'
						)
	args = parser.parse_args()
	return args

if __name__ == '__main__':
	args = argparser()
	start_ptr = []
	keywords_workbook, active_sheet, keywords = loadKeywordsIntoDict()
	wordVectors = []
	# getArticlesBBC(keywords)
	# try:
	# 	start_ptr = getArticlesCyberpost(keywords)
	# except:
	# 	pass
		
	# if start_ptr == []:
	# 	start_ptr = [len(keywords[k]) for k in keywords]

	# processArticle(keywords, start_ptr)
	# wordVectors = createDistanceSpreadsheet(args, keywords_workbook, active_sheet, keywords)
	visusalizeDistance(wordVectors, model_type=args.mode)